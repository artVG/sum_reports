import openpyxl
from datetime import datetime
from transaction import Transaction
from tn import TN
from pathlib import Path
from document import Document
from typing import Tuple, List
from reports import TransactionsOfContract


def replace_special(s: str) -> str:
    """replace forbidden characters with space"""
    result = s
    special = r'\/*?:[].'
    for letter in special:
        result = result.replace(letter, ' ')
    return result


def write_transaction(
        sheet,
        start_row: int,
        start_column: int,
        transaction: Transaction
) -> Tuple[int, int]:
    """write Transaction class data to excel shit
    starting at specified row and column
    :returns row and column next to the last of written"""
    sheet.cell(
        row=start_row,
        column=start_column,
        value=transaction.product
    )
    sheet.cell(
        row=start_row,
        column=start_column + 1,
        value=transaction.amount
    )
    sheet.cell(
        row=start_row,
        column=start_column + 2,
        value=transaction.price
    )
    next_row = start_row + 1
    next_column = start_column + 3
    return next_row, next_column


def write_document(
        sheet,
        start_row: int,
        start_column: int,
        document: Document
) -> Tuple[int, int]:
    """write Document class data to excel shit
    starting at specified row and column
    :returns row and column next to the last of written"""
    sheet.cell(
        row=start_row,
        column=start_column,
        value=document.series
    )
    sheet.cell(
        row=start_row,
        column=start_column + 1,
        value=document.number
    )
    sheet.cell(
        row=start_row,
        column=start_column + 2,
        value=document.date
    )
    next_row = start_row + 1
    next_column = start_column + 3
    return next_row, next_column


def write_tn(
        sheet,
        start_row: int,
        start_column: int,
        tn: TN
) -> Tuple[int, int]:
    """write TN class data to excel shit
    starting at specified row and column
    :returns row and column next to the last of written"""
    end: Tuple[int, int] = write_document(
        sheet=sheet,
        start_row=start_row,
        start_column=start_column,
        document=tn.document
    )
    sheet.cell(
        row=start_row,
        column=end[1],
        value=tn.contract
    )
    sheet.cell(
        row=start_row,
        column=end[1] + 1,
        value=tn.transactions_sum
    )
    for transaction in tn.transactions:
        end = write_transaction(
            sheet=sheet,
            start_row=end[0],
            start_column=start_column + 1,
            transaction=transaction
        )
    return end


def write_report_sorted_by_contract_sum_by_transaction(
        report: List[TransactionsOfContract],
        save_dir: Path
) -> None:
    # create excel file
    excel_file = openpyxl.Workbook()
    for report in report:
        # create new excel sheet for contract
        sheet = excel_file.create_sheet(replace_special(report.contract))
        # specify start cell
        start: Tuple[int, int] = (1, 1)
        for transaction in report.tns:
            # write data and get next to in row and column number
            start = write_transaction(
                sheet=sheet,
                start_row=start[0],
                start_column=1,
                transaction=transaction
            )
    # save excel file to specified folder and add date/time of creation to file name
    excel_file.save(
        filename=(save_dir / f'сортировка_по_контрактам_сумма_по_позициям_{datetime.now()}.xlsx'.replace(':', '-'))
    )


def write_report_sorted_by_contract(report: List[List[TN]], save_dir: Path) -> None:
    # create excel file
    excel_file = openpyxl.Workbook()
    for contract in report:
        # create new excel sheet for contract
        sheet = excel_file.create_sheet(replace_special(contract[0].contract))
        # specify start cell
        start: Tuple[int, int] = (1, 1)
        for tn in contract:
            # write data and get next to in row and column number
            start = write_tn(
                sheet=sheet,
                start_row=start[0],
                start_column=1,
                tn=tn
            )
    # save excel file to specified folder and add date/time of creation to file name
    excel_file.save(filename=(save_dir / f'сортировка_по_контрактам_{datetime.now()}.xlsx'.replace(':', '-')))


def write_report_sum_by_transaction(report: List[Transaction], save_dir: Path) -> None:
    # create excel file
    excel_file = openpyxl.Workbook()
    # get active excel sheet
    sheet = excel_file.active
    # specify start cell
    start: Tuple[int, int] = (1, 1)
    for transaction in report:
        # write data and get next to in row and column number
        start = write_transaction(
            sheet=sheet,
            start_row=start[0],
            start_column=1,
            transaction=transaction
        )
    # save excel file to specified folder and add date/time of creation to file name
    excel_file.save(filename=(save_dir / f'сумма_по_позициям_{datetime.now()}.xlsx'.replace(':', '-')))


def write_report_list_tn(report: List[TN], save_dir: Path) -> None:
    # create excel file
    excel_file = openpyxl.Workbook()
    # get active excel sheet
    sheet = excel_file.active
    # specify start cell
    start: Tuple[int, int] = (1, 1)
    for tn in report:
        # write data and get next to in row and column number
        start = write_tn(
            sheet=sheet,
            start_row=start[0],
            start_column=1,
            tn=tn
        )
    # save excel file to specified folder and add date/time of creation to file name
    excel_file.save(filename=(save_dir / f'список_накладных_{datetime.now()}.xlsx'.replace(':', '-')))
